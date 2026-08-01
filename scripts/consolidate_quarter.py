"""四半期/上期/下期ダッシュボード用 companies_<period>.json を一括生成。

サポート期間: q2 (2Q単独), h1 (上期), q3 (3Q単独), q3cum (3Q累計), q4 (4Q単独), h2 (下期)

データソース:
  当期/前期: 各 04/05/06/07/08/09_*_2026.03 Excel の最新ver
  前々期:    各 12/13/14/15/16/17_*_2024.03 Excel の最新ver

usage:
  python consolidate_quarter.py q2 h1 q3 q3cum q4 h2
  または引数なし → 全期間生成
"""
import json
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

from xlsx_utils import find_latest_xlsx, data_sheet, build_ltm, ltm_trend, margin

ROOT = Path(__file__).parent.parent.parent  # 01_通期実績_2026.03
PROJ = ROOT.parent  # 競合各社業績比較_20260520
OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# (period_key, label, ja_short, 当期Excelパターン, 前々期Excelパターン, current_period suffix(3月), suffix(8月))
PERIODS = {
    "q2": {
        "label": "2Q単独",
        "xlsx_curr": find_latest_xlsx(PROJ / "04_2Q単独_2026.03"),
        "xlsx_pp":   find_latest_xlsx(PROJ / "12_2Q単独_2024.03"),
        "curr_suffix_3": "-Q2", "curr_suffix_8": "-Q2",
        "prev_suffix_3": "-Q2", "prev_suffix_8": "-Q2",
        "pp_suffix_3":   "-Q2", "pp_suffix_8":   "-Q2",
    },
    "h1": {
        "label": "上期",
        "xlsx_curr": find_latest_xlsx(PROJ / "05_上期実績_2026.03"),
        "xlsx_pp":   find_latest_xlsx(PROJ / "13_上期実績_2024.03"),
        "curr_suffix_3": "-H1", "curr_suffix_8": "-H1",
        "prev_suffix_3": "-H1", "prev_suffix_8": "-H1",
        "pp_suffix_3":   "-H1", "pp_suffix_8":   "-H1",
    },
    "q3": {
        "label": "3Q単独",
        "xlsx_curr": find_latest_xlsx(PROJ / "06_3Q単独_2026.03"),
        "xlsx_pp":   find_latest_xlsx(PROJ / "14_3Q単独_2024.03"),
        "curr_suffix_3": "-Q3", "curr_suffix_8": "-Q3",
        "prev_suffix_3": "-Q3", "prev_suffix_8": "-Q3",
        "pp_suffix_3":   "-Q3", "pp_suffix_8":   "-Q3",
    },
    "q3cum": {
        "label": "3Q累計",
        "xlsx_curr": find_latest_xlsx(PROJ / "07_3Q累計_2026.03"),
        "xlsx_pp":   find_latest_xlsx(PROJ / "15_3Q累計_2024.03"),
        "curr_suffix_3": "-Q3CUM", "curr_suffix_8": "-Q3CUM",
        "prev_suffix_3": "-Q3CUM", "prev_suffix_8": "-Q3CUM",
        "pp_suffix_3":   "-Q3CUM", "pp_suffix_8":   "-Q3CUM",
    },
    "q4": {
        "label": "4Q単独",
        "xlsx_curr": find_latest_xlsx(PROJ / "08_4Q単独_2026.03"),
        "xlsx_pp":   find_latest_xlsx(PROJ / "16_4Q単独_2024.03"),
        "curr_suffix_3": "-Q4", "curr_suffix_8": "-Q4",
        "prev_suffix_3": "-Q4", "prev_suffix_8": "-Q4",
        "pp_suffix_3":   "-Q4", "pp_suffix_8":   "-Q4",
    },
    "h2": {
        "label": "下期",
        "xlsx_curr": find_latest_xlsx(PROJ / "09_下期実績_2026.03"),
        "xlsx_pp":   find_latest_xlsx(PROJ / "17_下期実績_2024.03"),
        "curr_suffix_3": "-H2", "curr_suffix_8": "-H2",
        "prev_suffix_3": "-H2", "prev_suffix_8": "-H2",
        "pp_suffix_3":   "-H2", "pp_suffix_8":   "-H2",
    },
}

# 各社の列レイアウト
COMPANIES_MAIN = [
    {"key": "yamada", "label": "ヤマダHD", "ticker": "9831", "fy_end": "03", "consol": "consolidated", "col_curr": 4, "col_prev": 5, "url": "https://www.yamada-holdings.jp/ir/"},
    {"key": "yamada_denki", "label": "ヤマダ（デンキセグメント）", "ticker": "9831-DK", "fy_end": "03", "consol": "segment", "col_curr": 6, "col_prev": 7, "url": "https://www.yamada-holdings.jp/ir/", "is_segment": True},
    {"key": "ks", "label": "ケーズHD", "ticker": "8282", "fy_end": "03", "consol": "consolidated", "col_curr": 8, "col_prev": 9, "url": "https://www.ksdenki.co.jp/ir/"},
    {"key": "edion", "label": "エディオン", "ticker": "2730", "fy_end": "03", "consol": "consolidated", "col_curr": 10, "col_prev": 11, "url": "https://www.edion.co.jp/ir/"},
    {"key": "joshin", "label": "上新電機", "ticker": "8173", "fy_end": "03", "consol": "consolidated", "col_curr": 12, "col_prev": 13, "url": "https://www.joshin.co.jp/ir/"},
    {"key": "nojima", "label": "ノジマ", "ticker": "7419", "fy_end": "03", "consol": "consolidated", "col_curr": 14, "col_prev": 15, "url": "https://www.nojima.co.jp/ir/"},
    {"key": "bic", "label": "ビックカメラ（連結）", "ticker": "3048", "fy_end": "08", "consol": "consolidated", "col_curr": 22, "col_prev": 23, "url": "https://www.biccamera.co.jp/ir/"},
    {"key": "kojima", "label": "コジマ", "ticker": "7513", "fy_end": "08", "consol": "non_consolidated", "col_curr": 18, "col_prev": 19, "url": "https://www.kojima.net/corporation/ir/"},
]

ROW_TO_METRIC = {
    7: "Revenue", 8: "GrossProfit", 9: "SGA",
    82: "OperatingIncome", 83: "OrdinaryIncome", 84: "NetIncome",
    88: "InterestBearingDebt", 89: "TotalEquity",
    100: "TotalAssets", 101: "Inventory", 86: "Tax",
}

# 通期計画（既存）— 末尾「次期通期業績予想」表用にforecast_annualで保持
FORECAST = {
    "yamada":       {"Revenue": 1780000, "OperatingIncome": 51500, "OrdinaryIncome": 52600, "NetIncome": 27800, "GrossProfit": 503100},
    "ks":           {"Revenue": 785000,  "OperatingIncome": 30500, "OrdinaryIncome": 33500, "NetIncome": 20000, "GrossProfit": 219500},
    "edion":        {"Revenue": 816000,  "OperatingIncome": 27000, "OrdinaryIncome": 27000, "NetIncome": 15700, "GrossProfit": 235800},
    "joshin":       {"Revenue": 438000,  "OperatingIncome": 6000,  "OrdinaryIncome": 5500,  "NetIncome": 3500,  "GrossProfit": 114500},
    "nojima":       {"Revenue": 1000000, "OperatingIncome": 59000, "OrdinaryIncome": 76000, "NetIncome": 48000, "GrossProfit": None},
    "bic":          {"Revenue": 1013000, "OperatingIncome": 30500, "OrdinaryIncome": 31500, "NetIncome": 17500, "GrossProfit": 271484},
    "kojima":       {"Revenue": 294000,  "OperatingIncome": 7600,  "OrdinaryIncome": 7900,  "NetIncome": 4900,  "GrossProfit": 79968},
    "yamada_denki": {"Revenue": 1407400, "OperatingIncome": 34500, "OrdinaryIncome": 37100, "NetIncome": None,  "GrossProfit": 411500},
}


def to_num(v):
    if v is None or (isinstance(v, str) and v.startswith("=")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def safe_div(a, b):
    try:
        return a / b if a is not None and b not in (None, 0) else None
    except (TypeError, ZeroDivisionError):
        return None


def build_one(period_key, conf):
    print(f"\n========== {period_key} ({conf['label']}) ==========")
    if not conf["xlsx_curr"].exists():
        print(f"  ERR xlsx_curr not found: {conf['xlsx_curr'].name}")
        return
    if not conf["xlsx_pp"].exists():
        print(f"  ERR xlsx_pp not found: {conf['xlsx_pp'].name}")
        return
    wb = load_workbook(conf["xlsx_curr"], data_only=False)
    ws = data_sheet(wb)
    wb_pp = load_workbook(conf["xlsx_pp"], data_only=True)
    ws_pp = data_sheet(wb_pp)

    companies_out = []
    for co in COMPANIES_MAIN:
        is_8 = (co["fy_end"] == "08")
        # 3月決算社: FY2026, 8月決算社: FY2025
        cp_base = "FY2025" if is_8 else "FY2026"
        pp_base = "FY2024" if is_8 else "FY2025"
        ppp_base = "FY2023" if is_8 else "FY2024"
        suf = conf["curr_suffix_8"] if is_8 else conf["curr_suffix_3"]
        sufp = conf["prev_suffix_8"] if is_8 else conf["prev_suffix_3"]
        sufpp = conf["pp_suffix_8"] if is_8 else conf["pp_suffix_3"]
        current_period = cp_base + suf
        previous_period = pp_base + sufp
        prev_previous_period = ppp_base + sufpp
        forecast_period = "FY2026" if is_8 else "FY2027"

        metrics = {}
        for row, key in ROW_TO_METRIC.items():
            cur = to_num(ws.cell(row=row, column=co["col_curr"]).value)
            prev = to_num(ws.cell(row=row, column=co["col_prev"]).value)
            prev_prev = to_num(ws_pp.cell(row=row, column=co["col_curr"]).value)
            fc_annual = FORECAST.get(co["key"], {}).get(key)
            metrics[key] = {
                "prev_previous": prev_prev,
                "current": cur,
                "previous": prev,
                "forecast": None,           # 四半期では推移グラフ4点目を描画しない
                "forecast_annual": fc_annual,
                "unit": "百万円",
            }

        # YoY
        yoy = {}
        for key in ("Revenue", "OperatingIncome", "OrdinaryIncome", "NetIncome"):
            m = metrics.get(key, {})
            cur = m.get("current"); prev = m.get("previous"); pp = m.get("prev_previous")
            yoy[key] = {
                "current_yoy_pct": round((cur / prev - 1) * 100, 2) if cur and prev else None,
                "previous_yoy_pct": round((prev / pp - 1) * 100, 2) if prev and pp else None,
                "forecast_yoy_pct": None,
            }

        # ratios
        rev = metrics["Revenue"]["current"]
        op = metrics["OperatingIncome"]["current"]
        ta = metrics["TotalAssets"]["current"]
        te = metrics["TotalEquity"]["current"]
        ord_c = metrics["OrdinaryIncome"]["current"]
        rev_p = metrics["Revenue"]["previous"]
        ord_p = metrics["OrdinaryIncome"]["previous"]
        rev_pp = metrics["Revenue"]["prev_previous"]
        ord_pp = metrics["OrdinaryIncome"]["prev_previous"]
        gp_c = metrics["GrossProfit"]["current"]
        gp_p = metrics["GrossProfit"]["previous"]
        gp_pp = metrics["GrossProfit"]["prev_previous"]
        ratios = {
            "operating_margin_pct":             round(op / rev * 100, 2) if op and rev else None,
            "ordinary_margin_pct":              round(ord_c / rev * 100, 2) if ord_c and rev else None,
            "ordinary_margin_pct_previous":     round(ord_p / rev_p * 100, 2) if ord_p and rev_p else None,
            "ordinary_margin_pct_prev_previous":round(ord_pp / rev_pp * 100, 2) if ord_pp and rev_pp else None,
            "equity_ratio_pct":                 round(te / ta * 100, 2) if te and ta else None,
            "gross_margin_pct":                 margin(gp_c, rev),
            "gross_margin_pct_previous":        margin(gp_p, rev_p),
            "gross_margin_pct_prev_previous":   margin(gp_pp, rev_pp),
            "gross_margin_pt_yoy": (round(margin(gp_c, rev) - margin(gp_p, rev_p), 2)
                                    if margin(gp_c, rev) is not None
                                    and margin(gp_p, rev_p) is not None else None),
            "net_margin_pct":                   margin(metrics["NetIncome"]["current"], rev),
            "net_margin_pct_previous":          margin(metrics["NetIncome"]["previous"], rev_p),
            "net_margin_pct_prev_previous":     margin(metrics["NetIncome"]["prev_previous"], rev_pp),
        }

        # 経常利益率の前期差(pt)、財務レバレッジ、ROE
        # 財務レバレッジ・ROEはExcelのR104(=総資産÷自己資本)・R98(=純利益÷自己資本×100)と同じ定義。
        # デンキセグメント等はBS非開示のため自己資本がなく、いずれもNoneになる。
        _om_c, _om_p = ratios["ordinary_margin_pct"], ratios["ordinary_margin_pct_previous"]
        ratios["ordinary_margin_pt_yoy"] = (round(_om_c - _om_p, 2)
                                            if _om_c is not None and _om_p is not None else None)
        ratios["financial_leverage"] = round(ta / te, 3) if ta and te else None
        # 純利益率の前期差(pt)
        _nm_c, _nm_p = ratios["net_margin_pct"], ratios["net_margin_pct_previous"]
        ratios["net_margin_pt_yoy"] = (round(_nm_c - _nm_p, 2)
                                       if _nm_c is not None and _nm_p is not None else None)
        # 財務レバレッジの前期差(倍)。前期末BSは metrics の previous から算出する
        _ta_p, _te_p = metrics["TotalAssets"]["previous"], metrics["TotalEquity"]["previous"]
        _lev_p = round(_ta_p / _te_p, 3) if _ta_p and _te_p else None
        ratios["financial_leverage_previous"] = _lev_p
        ratios["financial_leverage_diff"] = (round(ratios["financial_leverage"] - _lev_p, 3)
                                             if ratios["financial_leverage"] is not None
                                             and _lev_p is not None else None)
        ratios["roe_pct"] = margin(metrics["NetIncome"]["current"], te)

        # trend_ratios: BS依存の比率（ROE/ROIC/回転率）は四半期では推移グラフを描かないためnull。
        # 利益率（売上総利益率・純利益率）は期間按分の影響を受けないので四半期でも推移表示する。
        trend_ratios = {"roe": {}, "roic": {}, "asset_turnover": {}, "inventory_turnover": {},
                        "gross_margin": {}, "net_margin": {}}
        for key in trend_ratios.keys():
            trend_ratios[key] = {"prev_previous": None, "previous": None, "current": None, "forecast": None}
        for period in ("prev_previous", "previous", "current"):
            rev_p2 = metrics["Revenue"][period]
            trend_ratios["gross_margin"][period] = margin(metrics["GrossProfit"][period], rev_p2)
            trend_ratios["net_margin"][period] = margin(metrics["NetIncome"][period], rev_p2)

        if co.get("is_segment"):
            ratios["equity_ratio_pct"] = None

        # 直近四半期(LTM)回転率も推移グラフ用の形に持たせる
        ltm_data = build_ltm(ws, ws_pp, co)
        trend_ratios["ltm_asset_turnover"] = ltm_trend(ltm_data, "asset_turnover")
        trend_ratios["ltm_inventory_turnover"] = ltm_trend(ltm_data, "inventory_turnover")
        # ROE / ROIC は直近12ヶ月ベースに統一する。
        # 四半期の利益をそのまま自己資本で割ると期間の短さだけで小さく出るため。
        # 次期予想の点は従来どおり計画値ベースで残す（LTMは実績のみ算定できるため）
        _fc_roe = trend_ratios.get("roe", {}).get("forecast")
        _fc_roic = trend_ratios.get("roic", {}).get("forecast")
        trend_ratios["roe"] = ltm_trend(ltm_data, "roe_pct")
        trend_ratios["roic"] = ltm_trend(ltm_data, "roic_pct")
        trend_ratios["roe"]["forecast"] = _fc_roe
        trend_ratios["roic"]["forecast"] = _fc_roic
        ratios["roe_pct"] = ltm_data["current"]["roe_pct"]
        ratios["roic_pct"] = ltm_data["current"]["roic_pct"]
        # 直近12ヶ月ベース指標の前期差
        def _diff(cur, prev, nd):
            return round(cur - prev, nd) if cur is not None and prev is not None else None
        _lp, _lc = ltm_data["previous"], ltm_data["current"]
        ratios["roe_pt_yoy"] = _diff(_lc["roe_pct"], _lp["roe_pct"], 2)
        ratios["roic_pt_yoy"] = _diff(_lc["roic_pct"], _lp["roic_pct"], 2)
        ratios["asset_turnover_diff"] = _diff(_lc["asset_turnover"], _lp["asset_turnover"], 3)
        ratios["inventory_turnover_ltm"] = _lc["inventory_turnover"]
        ratios["inventory_turnover_diff"] = _diff(_lc["inventory_turnover"], _lp["inventory_turnover"], 3)

        companies_out.append({
            "key": co["key"], "label": co["label"], "ticker": co["ticker"],
            "fiscal_year_end": co["fy_end"], "consolidation": co["consol"],
            "current_period": current_period, "previous_period": previous_period,
            "prev_previous_period": prev_previous_period, "forecast_period": forecast_period,
            "tanshin_url": co["url"], "metrics": metrics, "yoy": yoy, "ratios": ratios,
            "trend_ratios": trend_ratios, "ltm": ltm_data,
            "is_segment": bool(co.get("is_segment")),
        })

    output = {
        "schema_version": "1.0",
        "generated_at": date.today().isoformat(),
        "source": f"各社決算短信・決算説明会資料 ({conf['label']})",
        "period_type": f"quarterly_{period_key}",
        "companies": companies_out,
    }
    out_path = OUT_DIR / f"companies_{period_key}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  → {out_path.name}")
    for c in companies_out:
        rev = c["metrics"]["Revenue"]["current"]
        op = c["metrics"]["OperatingIncome"]["current"]
        om = c["ratios"]["ordinary_margin_pct"]
        rev_s = f"{rev:,.0f}" if isinstance(rev,(int,float)) else "—"
        op_s = f"{op:,.0f}" if isinstance(op,(int,float)) else "—"
        om_s = f"{om}%" if om is not None else "—"
        print(f"    {c['label']:20s} ({c['current_period']:14s}): 売上 {rev_s:>11s} / 営利 {op_s:>8s} / 経常利益率 {om_s}")


def main():
    args = sys.argv[1:]
    if not args:
        args = list(PERIODS.keys())
    for pk in args:
        if pk not in PERIODS:
            print(f"  unknown period: {pk}")
            continue
        build_one(pk, PERIODS[pk])


if __name__ == "__main__":
    main()
