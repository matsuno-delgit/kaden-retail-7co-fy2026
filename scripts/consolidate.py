"""Excel ファイル(_ver.4.xlsx)から、ダッシュボード data/companies.json を生成する。

リファレンス index.html が期待するスキーマ:
{
  "schema_version": "1.0",
  "generated_at": "YYYY-MM-DD",
  "companies": [
    {
      "key": "yamada",
      "label": "ヤマダHD",
      "ticker": "9831",
      "fiscal_year_end": "03",
      "consolidation": "consolidated",
      "current_period": "FY2026",
      "previous_period": "FY2025",
      "forecast_period": "FY2027",
      "tanshin_url": "...",
      "metrics": {
        "Revenue": {"current":, "previous":, "forecast":, "unit": "百万円"},
        ...
      },
      "yoy": {"Revenue": {"current_yoy_pct":, "forecast_yoy_pct":}, ...},
      "ratios": {"operating_margin_pct":, "equity_ratio_pct":}
    },
    ...
  ]
}

メイン7社のみを対象とする（家電量販店7社業績比較）。
セグメント単位（デンキセグ・ノジマ家電専門店・ビックカメラ単体）は補足セクションに別途。
"""
import json
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

# プロジェクトルートからの相対パス
ROOT = Path(__file__).parent.parent.parent  # kaden-retail-7co-fy2026 の親
# 最新の通期実績 _ver.9 を採用（ROIC新式統一・法人税率参照式統一・不要数式削除済み）
XLSX = ROOT / "【経営企画部】各社業績対比フォーマット（2026.03通期）_エディオン2026.3期反映_20260521_ver.9.xlsx"
# 前々期実績 (2024年3月期通期、ビック/コジマは2023年8月期) を別Excelから読み込み
PREV_PREV_XLSX = ROOT.parent / "10_通期実績_2024.03" / "【経営企画部】各社業績対比フォーマット（2024.03通期）_ver.5.xlsx"
# 通期計画 (2027/3期計画、ビック/コジマは2026/8期計画) — 来期配当総額(N92等)を取得
PLAN_XLSX = ROOT.parent / "02_通期計画_2027.03" / "【経営企画部】各社業績対比フォーマット（2027.03通期計画 vs 2026.03通期実績）_ver.6.xlsx"
OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# 各社の定義（Excel列レイアウト + メタ）
# 表示順: ヤマダHD → ヤマダ（デンキセグメント） → ケーズHD → エディオン → 上新電機
#       → ノジマ → ビックカメラ（連結） → コジマ
COMPANIES_MAIN = [
    {
        "key": "yamada", "label": "ヤマダHD", "ticker": "9831",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026", "previous_period": "FY2025", "forecast_period": "FY2027",
        "col_curr": 4, "col_prev": 5,
        "tanshin_url": "https://www.yamada-holdings.jp/ir/",
    },
    # 補助エントリ: ヤマダHDの「デンキセグメント」を別建てで集計
    # BS関連（総資産・自己資本・有利子負債）はセグメント開示なしのためnull
    # 在庫回転率(R102)は説明会資料(ヤマダデンキPOSベース)の固定値で別途扱う
    {
        "key": "yamada_denki", "label": "ヤマダ（デンキセグメント）", "ticker": "9831-DK",
        "fiscal_year_end": "03", "consolidation": "segment",
        "current_period": "FY2026", "previous_period": "FY2025", "forecast_period": "FY2027",
        "col_curr": 6, "col_prev": 7,
        "tanshin_url": "https://www.yamada-holdings.jp/ir/",
        "is_segment": True,  # BS情報を持たない、ROE/ROIC/総資産回転率/自己資本比率 等を表示しない
    },
    {
        "key": "ks", "label": "ケーズHD", "ticker": "8282",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026", "previous_period": "FY2025", "forecast_period": "FY2027",
        "col_curr": 8, "col_prev": 9,
        "tanshin_url": "https://www.ksdenki.co.jp/ir/",
    },
    {
        "key": "edion", "label": "エディオン", "ticker": "2730",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026", "previous_period": "FY2025", "forecast_period": "FY2027",
        "col_curr": 10, "col_prev": 11,
        "tanshin_url": "https://www.edion.co.jp/ir/",
    },
    {
        "key": "joshin", "label": "上新電機", "ticker": "8173",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026", "previous_period": "FY2025", "forecast_period": "FY2027",
        "col_curr": 12, "col_prev": 13,
        "tanshin_url": "https://www.joshin.co.jp/ir/",
    },
    {
        "key": "nojima", "label": "ノジマ", "ticker": "7419",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026", "previous_period": "FY2025", "forecast_period": "FY2027",
        "col_curr": 14, "col_prev": 15,
        "tanshin_url": "https://www.nojima.co.jp/ir/",
    },
    {
        "key": "bic", "label": "ビックカメラ（連結）", "ticker": "3048",
        "fiscal_year_end": "08", "consolidation": "consolidated",
        "current_period": "FY2025", "previous_period": "FY2024", "forecast_period": "FY2026",
        "col_curr": 22, "col_prev": 23,
        "tanshin_url": "https://www.biccamera.co.jp/ir/",
    },
    {
        "key": "kojima", "label": "コジマ", "ticker": "7513",
        "fiscal_year_end": "08", "consolidation": "non_consolidated",
        "current_period": "FY2025", "previous_period": "FY2024", "forecast_period": "FY2026",
        "col_curr": 18, "col_prev": 19,
        "tanshin_url": "https://www.kojima.net/corporation/ir/",
    },
]

# Excel行 → メトリクスキー (リファレンスindex.htmlで使うキー名)
ROW_TO_METRIC = {
    7:  "Revenue",
    8:  "GrossProfit",
    9:  "SGA",
    82: "OperatingIncome",
    83: "OrdinaryIncome",
    84: "NetIncome",
    88: "InterestBearingDebt",
    89: "TotalEquity",         # 自己資本
    91: "DividendPerShare",
    92: "DividendTotal",
    93: "PayoutRatio",
    96: "EPS",
    100: "TotalAssets",
    101: "Inventory",
    86: "Tax",  # 通期：法人税（ROIC = (OP - Tax) / IC で利用）
}

# 業績予想（各社決算短信表紙＋決算説明会資料から抽出済）
# 3月決算5社=2027年3月期計画、Bic/Kojima=2026年8月期計画
# 出典詳細は 02_通期計画_2027.03/extracted/forecast_all.json
FORECAST = {
    "yamada": {"Revenue": 1780000, "OperatingIncome": 51500, "OrdinaryIncome": 52600, "NetIncome": 27800},
    "ks":     {"Revenue": 785000,  "OperatingIncome": 30500, "OrdinaryIncome": 33500, "NetIncome": 20000},
    "edion":  {"Revenue": 816000,  "OperatingIncome": 27000, "OrdinaryIncome": 27000, "NetIncome": 15700},
    "joshin": {"Revenue": 438000,  "OperatingIncome": 6000,  "OrdinaryIncome": 5500,  "NetIncome": 3500},
    "nojima": {"Revenue": 1000000, "OperatingIncome": 59000, "OrdinaryIncome": 76000, "NetIncome": 48000},
    "bic":    {"Revenue": 1013000, "OperatingIncome": 30500, "OrdinaryIncome": 31500, "NetIncome": 17500},
    "kojima": {"Revenue": 294000,  "OperatingIncome": 7600,  "OrdinaryIncome": 7900,  "NetIncome": 4900},
    # デンキセグメントは ヤマダ説明会資料202603 P36 連結計画 から
    "yamada_denki": {"Revenue": 1407400, "OperatingIncome": 34500, "OrdinaryIncome": 37100, "NetIncome": None},
}


def to_num(v):
    if v is None or (isinstance(v, str) and v.startswith("=")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    wb = load_workbook(XLSX, data_only=False)
    ws = wb["PL・BSデータ (2026.通期)"]
    # 前々期 (2024/3期 or 2023/8期) Excel
    wb_pp = load_workbook(PREV_PREV_XLSX, data_only=True)
    ws_pp = wb_pp.active
    # 通期計画 (来期配当総額の取得用)
    wb_plan = load_workbook(PLAN_XLSX, data_only=True)
    ws_plan = wb_plan.active

    companies_out = []
    for co in COMPANIES_MAIN:
        metrics = {}
        for row, key in ROW_TO_METRIC.items():
            cur = to_num(ws.cell(row=row, column=co["col_curr"]).value)
            prev = to_num(ws.cell(row=row, column=co["col_prev"]).value)
            # 前々期: 10_通期実績_2024.03 Excel の同列(=当期=2024/3期 or 2023/8期)
            prev_prev = to_num(ws_pp.cell(row=row, column=co["col_curr"]).value)
            fc = FORECAST.get(co["key"], {}).get(key)
            metrics[key] = {
                "prev_previous": prev_prev,
                "current": cur,
                "previous": prev,
                "forecast": fc,
                "unit": "百万円" if key not in ("DividendPerShare", "EPS", "PayoutRatio") else (
                    "円" if key in ("DividendPerShare", "EPS") else "%"
                ),
            }

        # 派生: YoY % (current vs previous), forecast YoY %
        yoy = {}
        for key in ("Revenue", "OperatingIncome", "OrdinaryIncome", "NetIncome"):
            m = metrics.get(key, {})
            cur = m.get("current")
            prev = m.get("previous")
            fc = m.get("forecast")
            yoy[key] = {
                "current_yoy_pct": round((cur / prev - 1) * 100, 2) if cur and prev else None,
                "forecast_yoy_pct": round((fc / cur - 1) * 100, 2) if fc and cur else None,
            }

        # 当期スポット比率（既存）
        rev = metrics["Revenue"]["current"]
        op = metrics["OperatingIncome"]["current"]
        ta = metrics["TotalAssets"]["current"]
        te = metrics["TotalEquity"]["current"]
        ratios = {
            "operating_margin_pct": round(op / rev * 100, 2) if op and rev else None,
            "equity_ratio_pct": round(te / ta * 100, 2) if te and ta else None,
            "gross_margin_pct": round(metrics["GrossProfit"]["current"] / rev * 100, 2)
                if metrics["GrossProfit"]["current"] and rev else None,
        }

        # 推移用比率（前期 / 当期 / 次期予想）
        # 次期予想のBSは開示されないため、当期BS（TotalAssets, TotalEquity, Inventory,
        # InterestBearingDebt）を暫定流用する。
        # ROIC = (OP - Tax) / (InterestBearingDebt + TotalEquity) × 100
        #   次期予想Taxは当期の OP→Tax 実効率 (Tax_curr / OP_curr) を OP_forecast に適用。
        def safe_div(a, b):
            try:
                return a / b if a is not None and b not in (None, 0) else None
            except (TypeError, ZeroDivisionError):
                return None

        def roe(ni, eq):
            v = safe_div(ni, eq)
            return round(v * 100, 2) if v is not None else None

        # ROIC = 営業利益 × (1 - 実効法人税率35%) / (有利子負債 + 自己資本) × 100
        # ※実効税率は固定35%。実際の法人税等(R86)は使わず、理論NOPATで算出。
        EFFECTIVE_TAX_RATE = 0.35

        def roic(op_v, _tax_v_unused, debt_v, eq_v):
            ic = (debt_v or 0) + (eq_v or 0)
            if not op_v or ic == 0:
                return None
            nopat = op_v * (1 - EFFECTIVE_TAX_RATE)
            return round(nopat / ic * 100, 2)

        def turnover(num, den):
            v = safe_div(num, den)
            return round(v, 3) if v is not None else None

        # 当期実効率（Tax / OP） — 次期予想のTax推計に使う
        op_curr = metrics["OperatingIncome"]["current"]
        tax_curr = metrics.get("Tax", {}).get("current")
        tax_curr_rate = safe_div(tax_curr, op_curr) if op_curr and tax_curr else None

        # ---- 来期計画用：期末予想自己資本＝当期末自己資本＋来期純利益−来期配当総額 ----
        # （Excel 02_通期計画の式 N89=O89+N84-N92 と整合）
        n92 = to_num(ws_plan.cell(row=92, column=co["col_curr"]).value)  # 計画配当総額
        if n92 is None:
            # 数式 N92 = O92 × N91 / O91 で再計算（ExcelのDPS比例配当総額推定式）
            o92 = to_num(ws_plan.cell(row=92, column=co["col_prev"]).value)
            n91 = to_num(ws_plan.cell(row=91, column=co["col_curr"]).value)
            o91 = to_num(ws_plan.cell(row=91, column=co["col_prev"]).value)
            if o92 is not None and n91 is not None and o91 is not None and o91 != 0:
                n92 = o92 * n91 / o91
            else:
                # 推定不可（例: ノジマはO91クリア）→ 前期配当総額(O92)で代用
                n92 = o92
        forecast_dividend = n92
        ni_fc = metrics["NetIncome"]["forecast"]
        te_curr = metrics["TotalEquity"]["current"]
        ta_curr = metrics["TotalAssets"]["current"]
        # 期末予想自己資本 = 当期末te + 来期純利益 − 来期配当総額
        forecast_te = None
        if te_curr is not None and ni_fc is not None:
            forecast_te = te_curr + ni_fc - (forecast_dividend or 0)
        # 期末予想総資産 = 当期末ta + 来期純利益 − 来期配当総額 (Excel D100=E100+D84-E92)
        forecast_ta = None
        if ta_curr is not None and ni_fc is not None:
            forecast_ta = ta_curr + ni_fc - (forecast_dividend or 0)

        # 各期 BS
        bs_periods = {
            "prev_previous": {
                "ta": metrics["TotalAssets"]["prev_previous"],
                "te": metrics["TotalEquity"]["prev_previous"],
                "inv": metrics["Inventory"]["prev_previous"],
                "debt": metrics["InterestBearingDebt"]["prev_previous"],
            },
            "previous": {
                "ta": metrics["TotalAssets"]["previous"],
                "te": metrics["TotalEquity"]["previous"],
                "inv": metrics["Inventory"]["previous"],
                "debt": metrics["InterestBearingDebt"]["previous"],
            },
            "current": {
                "ta": metrics["TotalAssets"]["current"],
                "te": metrics["TotalEquity"]["current"],
                "inv": metrics["Inventory"]["current"],
                "debt": metrics["InterestBearingDebt"]["current"],
            },
            "forecast": {  # 期末予想BS = 当期末BS+来期利益-来期配当 (Excel式と整合)
                "ta": forecast_ta if forecast_ta is not None else metrics["TotalAssets"]["current"],
                "te": forecast_te if forecast_te is not None else metrics["TotalEquity"]["current"],
                "inv": metrics["Inventory"]["current"],            # 商品はBS非開示=当期流用
                "debt": metrics["InterestBearingDebt"]["current"], # 有利子負債もExcel D88=E88で当期末流用
                "is_provisional": True,
                "forecast_dividend_used": forecast_dividend,
            },
        }

        trend_ratios = {"roe": {}, "roic": {}, "asset_turnover": {}, "inventory_turnover": {}}
        for period in ("prev_previous", "previous", "current", "forecast"):
            rev_p = metrics["Revenue"][period]
            op_p = metrics["OperatingIncome"][period]
            ni_p = metrics["NetIncome"][period]
            bs = bs_periods[period]

            # Tax: 当期/前期は実額、forecast は当期実効率で推計
            if period == "forecast":
                tax_p = op_p * tax_curr_rate if op_p and tax_curr_rate else None
            else:
                tax_p = metrics.get("Tax", {}).get(period)

            trend_ratios["roe"][period] = roe(ni_p, bs["te"])
            trend_ratios["roic"][period] = roic(op_p, tax_p, bs["debt"], bs["te"])
            trend_ratios["asset_turnover"][period] = turnover(rev_p, bs["ta"])
            trend_ratios["inventory_turnover"][period] = turnover(rev_p, bs["inv"])

        # 暫定フラグ（次期BSが当期流用であることを明示）
        trend_ratios["forecast_uses_current_bs"] = True

        # ----- ヤマダHD連結 の在庫回転率 -----
        # ExcelのD102/E102と同じ計算ロジック (=Revenue/Inventory) で算出する。
        # InventoryはR101「商品及び製品+販売用不動産」（住建セグの住宅在庫を含む）。
        # （以前は説明会資料P28のヤマダデンキPOSベース固定値 4.0/4.5/5.0 を採用していたが、
        #   ユーザー指示によりExcel D102/E102 と同じ自然計算に変更。
        #   POSベース値は ヤマダ（デンキセグメント）の方で別建てに表示。）

        # ----- デンキセグメントは BS情報なし → ROE/ROIC/総資産回転率/自己資本比率 を null 化 -----
        # 在庫回転率のみ ヤマダデンキPOSベース固定値で別建て表示
        if co.get("is_segment"):
            for k in ("roe", "roic", "asset_turnover"):
                trend_ratios[k] = {"prev_previous": None, "previous": None, "current": None, "forecast": None}
            # 在庫回転率は POS ベース固定値 (ヤマダデンキ実績/計画)
            # 前々期(2024/3期)=3.6 / 前期(2025/3期)=4.0 / 当期(2026/3期)=4.5 / 計画(2027/3期)=5.0
            trend_ratios["inventory_turnover"] = {
                "prev_previous": 3.6, "previous": 4.0, "current": 4.5, "forecast": 5.0,
            }
            trend_ratios["inventory_turnover_override"] = {
                "basis": "デンキセグメント（ヤマダデンキPOSベース）",
                "source": "ヤマダ_決算説明会資料20260508 P28・P44 (前々期は2024/3期通期実績 POSベース 3.6)",
                "note": "セグメント実績(POSベース)。連結BSベースの計算ではなく説明会資料記載値",
            }
            # ratios の equity_ratio_pct と asset_turnover も null化
            ratios["equity_ratio_pct"] = None

        # 前々期年度ラベル
        prev_prev_period = {
            "FY2026": "FY2024", "FY2025": "FY2023",  # 通期: 3月決算社 / 8月決算社
        }.get(co["current_period"], None)

        co_out = {
            "key": co["key"],
            "label": co["label"],
            "ticker": co["ticker"],
            "fiscal_year_end": co["fiscal_year_end"],
            "consolidation": co["consolidation"],
            "current_period": co["current_period"],
            "previous_period": co["previous_period"],
            "prev_previous_period": prev_prev_period,
            "forecast_period": co["forecast_period"],
            "tanshin_url": co["tanshin_url"],
            "metrics": metrics,
            "yoy": yoy,
            "ratios": ratios,
            "trend_ratios": trend_ratios,
            "is_segment": bool(co.get("is_segment")),  # ヤマダ（デンキセグ）等のフラグ
        }
        companies_out.append(co_out)

    output = {
        "schema_version": "1.0",
        "generated_at": date.today().isoformat(),
        "source": "各社決算短信・決算説明会資料 (TDnet公開資料、2026/5月発表分)",
        "note": "ヤマダ・ケーズ・エディオン・上新・ノジマ=2026年3月期通期。ビックカメラ・コジマ=2025年8月期通期。コジマは非連結開示。",
        "companies": companies_out,
    }

    out_path = OUT_DIR / "companies.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path}")
    print(f"Companies: {len(companies_out)}")
    for c in companies_out:
        rev = c["metrics"]["Revenue"]["current"]
        op = c["metrics"]["OperatingIncome"]["current"]
        om = c["ratios"]["operating_margin_pct"]
        print(f"  {c['label']:10s} ({c['current_period']}): 売上 {rev:>12,.0f} / 営利 {op:>8,.0f} / 営利率 {om}%")

    # trend_ratios summary
    print("\n=== 推移比率（前期 / 当期 / 次期予想） ===")
    print(f"{'会社':10s}  {'ROE (%)':>23s}  {'ROIC (%)':>23s}  {'資産回転(回)':>23s}  {'在庫回転(回)':>23s}")
    for c in companies_out:
        tr = c["trend_ratios"]
        def f(d, key, fmt):
            v = d.get(key)
            return (fmt % v) if v is not None else "—"
        roe_p = f(tr["roe"], "previous", "%6.2f")
        roe_c = f(tr["roe"], "current",  "%6.2f")
        roe_f = f(tr["roe"], "forecast", "%6.2f")
        roic_p = f(tr["roic"], "previous", "%6.2f")
        roic_c = f(tr["roic"], "current",  "%6.2f")
        roic_f = f(tr["roic"], "forecast", "%6.2f")
        at_p = f(tr["asset_turnover"], "previous", "%5.3f")
        at_c = f(tr["asset_turnover"], "current",  "%5.3f")
        at_f = f(tr["asset_turnover"], "forecast", "%5.3f")
        it_p = f(tr["inventory_turnover"], "previous", "%5.2f")
        it_c = f(tr["inventory_turnover"], "current",  "%5.2f")
        it_f = f(tr["inventory_turnover"], "forecast", "%5.2f")
        print(
            f"  {c['label']:10s}  {roe_p}/{roe_c}/{roe_f}    "
            f"{roic_p}/{roic_c}/{roic_f}    "
            f"{at_p}/{at_c}/{at_f}    "
            f"{it_p}/{it_c}/{it_f}"
        )


if __name__ == "__main__":
    main()
