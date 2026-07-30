"""組替版ダッシュボード用 companies_kumikae_<period>.json を一括生成。

入力Excel (期間揃え比較フォルダ):
  当期版    1_組替通期_2026.02 etc.    (1_-8_)
  前々期版  9_組替通期_2025.02 etc.   (9_-16_)

期間ラベル統一: 全社「FY2026 (=2025/3〜2026/2 系)」「FY2025」「FY2024」を採用。
forecast/forecast_annual は組替版では使わない (null)。
"""
import json
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

from xlsx_utils import find_latest_xlsx, data_sheet, build_ltm, ltm_trend, margin

ROOT = Path(__file__).parent.parent.parent  # 01_通期実績_2026.03
PROJ = ROOT.parent  # 競合各社業績比較_20260520
KUMIKAE = PROJ / "期間揃え比較"
OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# 期間別 (curr_xlsx, pp_xlsx) ペア
PERIODS = {
    "fy": {
        "label": "組替通期 (2025/3〜2026/2)",
        "curr": find_latest_xlsx(KUMIKAE / "01_組替通期_2026.02"),
        "pp":   find_latest_xlsx(KUMIKAE / "09_組替通期_2025.02"),
    },
    "q1": {
        "label": "組替1Q (3〜5月)",
        "curr": find_latest_xlsx(KUMIKAE / "02_組替1Q_2025.05"),
        "pp":   find_latest_xlsx(KUMIKAE / "10_組替1Q_2024.05"),
    },
    "q2": {
        "label": "組替2Q単独 (6〜8月)",
        "curr": find_latest_xlsx(KUMIKAE / "03_組替2Q単独_2025.08"),
        "pp":   find_latest_xlsx(KUMIKAE / "11_組替2Q単独_2024.08"),
    },
    "h1": {
        "label": "組替上期 (3〜8月)",
        "curr": find_latest_xlsx(KUMIKAE / "04_組替上期_2025.08"),
        "pp":   find_latest_xlsx(KUMIKAE / "12_組替上期_2024.08"),
    },
    "q3": {
        "label": "組替3Q単独 (9〜11月)",
        "curr": find_latest_xlsx(KUMIKAE / "05_組替3Q単独_2025.11"),
        "pp":   find_latest_xlsx(KUMIKAE / "13_組替3Q単独_2024.11"),
    },
    "q3cum": {
        "label": "組替3Q累計 (3〜11月)",
        "curr": find_latest_xlsx(KUMIKAE / "06_組替3Q累計_2025.11"),
        "pp":   find_latest_xlsx(KUMIKAE / "14_組替3Q累計_2024.11"),
    },
    "q4": {
        "label": "組替4Q単独 (12〜翌2月)",
        "curr": find_latest_xlsx(KUMIKAE / "07_組替4Q単独_2026.02"),
        "pp":   find_latest_xlsx(KUMIKAE / "15_組替4Q単独_2025.02"),
    },
    "h2": {
        "label": "組替下期 (9〜翌2月)",
        "curr": find_latest_xlsx(KUMIKAE / "08_組替下期_2026.02"),
        "pp":   find_latest_xlsx(KUMIKAE / "16_組替下期_2025.02"),
    },
}

COMPANIES_MAIN = [
    {"key": "yamada", "label": "ヤマダHD", "ticker": "9831", "fy_end": "03", "consol": "consolidated", "col_curr": 4, "col_prev": 5, "url": "https://www.yamada-holdings.jp/ir/"},
    {"key": "yamada_denki", "label": "ヤマダ（デンキセグメント）", "ticker": "9831-DK", "fy_end": "03", "consol": "segment", "col_curr": 6, "col_prev": 7, "url": "https://www.yamada-holdings.jp/ir/", "is_segment": True},
    {"key": "ks", "label": "ケーズHD", "ticker": "8282", "fy_end": "03", "consol": "consolidated", "col_curr": 8, "col_prev": 9, "url": "https://www.ksdenki.co.jp/ir/"},
    {"key": "edion", "label": "エディオン", "ticker": "2730", "fy_end": "03", "consol": "consolidated", "col_curr": 10, "col_prev": 11, "url": "https://www.edion.co.jp/ir/"},
    {"key": "joshin", "label": "上新電機", "ticker": "8173", "fy_end": "03", "consol": "consolidated", "col_curr": 12, "col_prev": 13, "url": "https://www.joshin.co.jp/ir/"},
    {"key": "nojima", "label": "ノジマ", "ticker": "7419", "fy_end": "03", "consol": "consolidated", "col_curr": 14, "col_prev": 15, "url": "https://www.nojima.co.jp/ir/"},
    {"key": "bic", "label": "ビックカメラ（連結）", "ticker": "3048", "fy_end": "08", "consol": "consolidated", "col_curr": 22, "col_prev": 23, "url": "https://www.biccamera.co.jp/ir/"},
    {"key": "kojima", "label": "コジマ", "ticker": "7513", "fy_end": "08", "consol": "non_consolidated", "col_curr": 18, "col_prev": 19, "url": "https://www.kojima.net/corporation/ir/"},
]

ROW_TO_METRIC = {
    7: "Revenue", 8: "GrossProfit", 9: "SGA",
    82: "OperatingIncome", 83: "OrdinaryIncome", 84: "NetIncome",
    88: "InterestBearingDebt", 89: "TotalEquity",
    100: "TotalAssets", 101: "Inventory", 86: "Tax",
}


def to_num(v):
    if v is None or (isinstance(v, str) and v.startswith("=")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def safe_div(a, b):
    try:
        return a / b if a is not None and b not in (None, 0) else None
    except (TypeError, ZeroDivisionError):
        return None


def build_one(pk, conf, is_annual):
    print(f"\n========== {pk} ({conf['label']}) ==========")
    wb_curr = load_workbook(conf["curr"], data_only=True); ws_curr = data_sheet(wb_curr)
    wb_pp = load_workbook(conf["pp"], data_only=True); ws_pp = data_sheet(wb_pp)

    companies_out = []
    for co in COMPANIES_MAIN:
        # 統一期間ラベル: 全社FY2026/FY2025/FY2024
        current_period = "FY2026"
        previous_period = "FY2025"
        prev_previous_period = "FY2024"
        forecast_period = "FY2027"

        metrics = {}
        for row, key in ROW_TO_METRIC.items():
            cur = to_num(ws_curr.cell(row=row, column=co["col_curr"]).value)
            prev = to_num(ws_curr.cell(row=row, column=co["col_prev"]).value)
            # 前々期: 9_-16_の前期列 (= 2023/3-2024/2 期)
            prev_prev = to_num(ws_pp.cell(row=row, column=co["col_prev"]).value)
            metrics[key] = {
                "prev_previous": prev_prev,
                "current": cur,
                "previous": prev,
                "forecast": None,
                "forecast_annual": None,
                "unit": "百万円",
            }

        yoy = {}
        for key in ("Revenue", "OperatingIncome", "OrdinaryIncome", "NetIncome"):
            m = metrics.get(key, {})
            cur = m.get("current"); prev = m.get("previous"); pp = m.get("prev_previous")
            yoy[key] = {
                "current_yoy_pct": round((cur / prev - 1) * 100, 2) if cur and prev else None,
                "previous_yoy_pct": round((prev / pp - 1) * 100, 2) if prev and pp else None,
                "forecast_yoy_pct": None,
            }

        rev = metrics["Revenue"]["current"]
        op = metrics["OperatingIncome"]["current"]
        ta = metrics["TotalAssets"]["current"]
        te = metrics["TotalEquity"]["current"]
        ord_c = metrics["OrdinaryIncome"]["current"]
        rev_p = metrics["Revenue"]["previous"]
        ord_p = metrics["OrdinaryIncome"]["previous"]
        rev_pp = metrics["Revenue"]["prev_previous"]
        ord_pp = metrics["OrdinaryIncome"]["prev_previous"]
        gp_c = metrics["GrossProfit"]["current"]
        gp_p = metrics["GrossProfit"]["previous"]
        gp_pp = metrics["GrossProfit"]["prev_previous"]
        ratios = {
            "operating_margin_pct":              round(op / rev * 100, 2) if op and rev else None,
            "ordinary_margin_pct":               round(ord_c / rev * 100, 2) if ord_c and rev else None,
            "ordinary_margin_pct_previous":      round(ord_p / rev_p * 100, 2) if ord_p and rev_p else None,
            "ordinary_margin_pct_prev_previous": round(ord_pp / rev_pp * 100, 2) if ord_pp and rev_pp else None,
            "equity_ratio_pct":                  round(te / ta * 100, 2) if te and ta else None,
            "gross_margin_pct":                  margin(gp_c, rev),
            "gross_margin_pct_previous":         margin(gp_p, rev_p),
            "gross_margin_pct_prev_previous":    margin(gp_pp, rev_pp),
            "gross_margin_pt_yoy": (round(margin(gp_c, rev) - margin(gp_p, rev_p), 2)
                                    if margin(gp_c, rev) is not None
                                    and margin(gp_p, rev_p) is not None else None),
            "net_margin_pct":                    margin(metrics["NetIncome"]["current"], rev),
            "net_margin_pct_previous":           margin(metrics["NetIncome"]["previous"], rev_p),
            "net_margin_pct_prev_previous":      margin(metrics["NetIncome"]["prev_previous"], rev_pp),
        }

        # 経常利益率の前期差(pt)、財務レバレッジ、ROE
        # 財務レバレッジ・ROEはExcelのR104(=総資産÷自己資本)・R98(=純利益÷自己資本×100)と同じ定義。
        # デンキセグメント等はBS非開示のため自己資本がなく、いずれもNoneになる。
        _om_c, _om_p = ratios["ordinary_margin_pct"], ratios["ordinary_margin_pct_previous"]
        ratios["ordinary_margin_pt_yoy"] = (round(_om_c - _om_p, 2)
                                            if _om_c is not None and _om_p is not None else None)
        ratios["financial_leverage"] = round(ta / te, 3) if ta and te else None
        ratios["roe_pct"] = margin(metrics["NetIncome"]["current"], te)

        # trend_ratios (通期のみ計算、その他はnull)
        EFFECTIVE_TAX_RATE = 0.35

        def roe(ni, eq):
            v = safe_div(ni, eq)
            return round(v * 100, 2) if v is not None else None

        def roic(op_v, debt_v, eq_v):
            ic = (debt_v or 0) + (eq_v or 0)
            if not op_v or ic == 0:
                return None
            return round(op_v * (1 - EFFECTIVE_TAX_RATE) / ic * 100, 2)

        def turnover(num, den):
            v = safe_div(num, den)
            return round(v, 3) if v is not None else None

        trend_ratios = {"roe": {}, "roic": {}, "asset_turnover": {}, "inventory_turnover": {},
                        "gross_margin": {}, "net_margin": {}}
        if is_annual:
            bs_periods = {
                "prev_previous": {"ta": metrics["TotalAssets"]["prev_previous"], "te": metrics["TotalEquity"]["prev_previous"], "inv": metrics["Inventory"]["prev_previous"], "debt": metrics["InterestBearingDebt"]["prev_previous"]},
                "previous":      {"ta": metrics["TotalAssets"]["previous"],      "te": metrics["TotalEquity"]["previous"],      "inv": metrics["Inventory"]["previous"],      "debt": metrics["InterestBearingDebt"]["previous"]},
                "current":       {"ta": metrics["TotalAssets"]["current"],       "te": metrics["TotalEquity"]["current"],       "inv": metrics["Inventory"]["current"],       "debt": metrics["InterestBearingDebt"]["current"]},
            }
            for period in ("prev_previous", "previous", "current"):
                rev_p2 = metrics["Revenue"][period]
                op_p2 = metrics["OperatingIncome"][period]
                ni_p2 = metrics["NetIncome"][period]
                bs = bs_periods[period]
                trend_ratios["roe"][period] = roe(ni_p2, bs["te"])
                trend_ratios["roic"][period] = roic(op_p2, bs["debt"], bs["te"])
                trend_ratios["asset_turnover"][period] = turnover(rev_p2, bs["ta"])
                trend_ratios["inventory_turnover"][period] = turnover(rev_p2, bs["inv"])
            for k in trend_ratios.keys():
                trend_ratios[k]["forecast"] = None
        else:
            for k in trend_ratios.keys():
                trend_ratios[k] = {"prev_previous": None, "previous": None, "current": None, "forecast": None}
        # 利益率は期間按分の影響を受けないため、四半期でも推移を出す
        for period in ("prev_previous", "previous", "current"):
            rev_p3 = metrics["Revenue"][period]
            trend_ratios["gross_margin"][period] = margin(metrics["GrossProfit"][period], rev_p3)
            trend_ratios["net_margin"][period] = margin(metrics["NetIncome"][period], rev_p3)

        # デンキセグ: BS非開示。在庫回転率は会社開示値（ヤマダデンキPOSベース）を採用。
        # 組替通期の期間は3月決算社にとって通常の通期と同じため開示値をそのまま使える。
        # 前々期(2024/3期)=3.65 は在庫回転日数100日の開示より 365÷100 で換算。
        if co.get("is_segment"):
            for k in ("roe", "roic", "asset_turnover"):
                trend_ratios[k] = {"prev_previous": None, "previous": None, "current": None, "forecast": None}
            if is_annual:
                trend_ratios["inventory_turnover"] = {
                    "prev_previous": 3.65, "previous": 4.0, "current": 4.5, "forecast": None,
                }
            ratios["equity_ratio_pct"] = None

        # 直近四半期(LTM)回転率も推移グラフ用の形に持たせる
        ltm_data = build_ltm(ws_curr, ws_pp, co, pp_col_key="col_prev")
        trend_ratios["ltm_asset_turnover"] = ltm_trend(ltm_data, "asset_turnover")
        trend_ratios["ltm_inventory_turnover"] = ltm_trend(ltm_data, "inventory_turnover")
        # ROE / ROIC は直近12ヶ月ベースに統一する。
        # 四半期の利益をそのまま自己資本で割ると期間の短さだけで小さく出るため。
        # 次期予想の点は従来どおり計画値ベースで残す（LTMは実績のみ算定できるため）
        _fc_roe = trend_ratios.get("roe", {}).get("forecast")
        _fc_roic = trend_ratios.get("roic", {}).get("forecast")
        trend_ratios["roe"] = ltm_trend(ltm_data, "roe_pct")
        trend_ratios["roic"] = ltm_trend(ltm_data, "roic_pct")
        trend_ratios["roe"]["forecast"] = _fc_roe
        trend_ratios["roic"]["forecast"] = _fc_roic
        ratios["roe_pct"] = ltm_data["current"]["roe_pct"]
        ratios["roic_pct"] = ltm_data["current"]["roic_pct"]

        companies_out.append({
            "key": co["key"], "label": co["label"], "ticker": co["ticker"],
            "fiscal_year_end": co["fy_end"], "consolidation": co["consol"],
            "current_period": current_period, "previous_period": previous_period,
            "prev_previous_period": prev_previous_period, "forecast_period": forecast_period,
            "tanshin_url": co["url"], "metrics": metrics, "yoy": yoy, "ratios": ratios,
            "trend_ratios": trend_ratios,
            "ltm": ltm_data,
            "is_segment": bool(co.get("is_segment")),
        })

    output = {
        "schema_version": "1.0",
        "generated_at": date.today().isoformat(),
        "source": f"組替版 ({conf['label']})",
        "period_type": f"kumikae_{pk}",
        "dataset_type": "kumikae",
        "companies": companies_out,
    }
    out_path = OUT_DIR / f"companies_kumikae_{pk}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"  → {out_path.name}")
    for c in companies_out:
        rev = c["metrics"]["Revenue"]["current"]
        ord_v = c["metrics"]["OrdinaryIncome"]["current"]
        om = c["ratios"]["ordinary_margin_pct"]
        rev_s = f"{rev:,.0f}" if isinstance(rev,(int,float)) else "—"
        ord_s = f"{ord_v:,.0f}" if isinstance(ord_v,(int,float)) else "—"
        om_s = f"{om}%" if om is not None else "—"
        print(f"    {c['label']:20s}: 売上={rev_s:>11s} / 経常={ord_s:>8s} / 経常率={om_s}")


def main():
    for pk, conf in PERIODS.items():
        build_one(pk, conf, is_annual=(pk == "fy"))


if __name__ == "__main__":
    main()
