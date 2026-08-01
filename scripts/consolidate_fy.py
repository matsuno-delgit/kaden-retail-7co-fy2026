"""会計年度セレクタ用の companies_fy{YYYY}_{period}.json を生成する。

既存の consolidate*.py が「2026年3月期を当期」とする前提で組まれているのに対し、
本スクリプトは任意の会計年度を「当期」として出力する。

  fy2027          … 20_1Q実績_2027.03_通常版（1Qのみ。ノジマ以外の3月決算社は未開示）
  fy2027_kumikae  … 19_1Q実績_2027.03（同上・ビック/コジマを暦月で期間揃え）
  fy2025          … 21_〜28_ の2025.03系列（8期間フル）
  fy2025_kumikae  … 期間揃え比較 09_〜16_（8期間フル）
  fy2024          … 10_〜17_ の2024.03系列（8期間フル）

fy2026 は既存の companies*.json をそのまま使うため対象外。

usage: python consolidate_fy.py
"""
import json
import sys
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

from xlsx_utils import (find_latest_xlsx, data_sheet, build_ltm, read_ltm, ltm_trend,
                        margin, period_label, kumikae_labels_08)

ROOT = Path(__file__).parent.parent.parent          # 01_通期実績_2026.03
PROJ = ROOT.parent                                   # 競合各社業績比較_20260520
OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

COMPANIES_MAIN = [
    {"key": "yamada", "label": "ヤマダHD", "ticker": "9831", "fy_end": "03", "consol": "consolidated", "col_curr": 4, "col_prev": 5, "url": "https://www.yamada-holdings.jp/ir/"},
    {"key": "yamada_denki", "label": "ヤマダ（デンキセグメント）", "ticker": "9831-DK", "fy_end": "03", "consol": "segment", "col_curr": 6, "col_prev": 7, "url": "https://www.yamada-holdings.jp/ir/", "is_segment": True},
    {"key": "ks", "label": "ケーズHD", "ticker": "8282", "fy_end": "03", "consol": "consolidated", "col_curr": 8, "col_prev": 9, "url": "https://www.ksdenki.co.jp/ir/"},
    {"key": "edion", "label": "エディオン", "ticker": "2730", "fy_end": "03", "consol": "consolidated", "col_curr": 10, "col_prev": 11, "url": "https://www.edion.co.jp/ir/"},
    {"key": "joshin", "label": "上新電機", "ticker": "8173", "fy_end": "03", "consol": "consolidated", "col_curr": 12, "col_prev": 13, "url": "https://www.joshin.co.jp/ir/"},
    {"key": "nojima", "label": "ノジマ", "ticker": "7419", "fy_end": "03", "consol": "consolidated", "col_curr": 14, "col_prev": 15, "url": "https://www.nojima.co.jp/ir/"},
    {"key": "bic", "label": "ビックカメラ（連結）", "ticker": "3048", "fy_end": "08", "consol": "consolidated", "col_curr": 22, "col_prev": 23, "url": "https://www.biccamera.co.jp/ir/"},
    {"key": "kojima", "label": "コジマ", "ticker": "7513", "fy_end": "08", "consol": "non_consolidated", "col_curr": 18, "col_prev": 19, "url": "https://www.kojima.net/corporation/ir/"},
]

ROW_TO_METRIC = {
    7: "Revenue", 8: "GrossProfit", 9: "SGA",
    82: "OperatingIncome", 83: "OrdinaryIncome", 84: "NetIncome",
    88: "InterestBearingDebt", 89: "TotalEquity",
    100: "TotalAssets", 101: "Inventory", 86: "Tax",
}

# 会計年度 → {period: (当期Excelフォルダ, 前々期の取得元)}
# 前々期の取得元は {決算月: (Excelフォルダ, 参照する列)} 。None なら推移グラフの1点目が空欄。
# 8月決算社は当期の対応づけが年度ごとに変わるため、決算月ごとに指定できるようにしている。
FISCAL_YEARS = {
    # 通常版の2027年3月期。8月決算社は四半期番号どうしで突き合わせる
    # （3月決算社の1Q＝2026/4-6 ↔ ビック・コジマの2026年8月期1Q＝2025/9-11）。
    "fy2027": {
        "label": "2027年3月期",
        "curr_fy": {"03": "FY2027", "08": "FY2026"},
        "prev_fy": {"03": "FY2026", "08": "FY2025"},
        "pp_fy":   {"03": "FY2025", "08": "FY2024"},
        "periods": {"q1": ("20_1Q実績_2027.03_通常版",
                           {"03": ("03_1Q実績_2026.03", "col_prev"),
                            "08": ("03_1Q実績_2026.03", "col_prev")})},
    },
    # 組替版の2027年3月期。ビック・コジマは暦月を合わせて2026年8月期3Q単独(2026/3-5)。
    "fy2027_kumikae": {
        "label": "2027年3月期（組替）",
        "kumikae_base": 2027,
        "curr_fy": {"03": "FY2027", "08": "FY2026"},
        "prev_fy": {"03": "FY2026", "08": "FY2025"},
        "pp_fy":   {"03": "FY2025", "08": "FY2024"},
        "periods": {"q1": ("19_1Q実績_2027.03",
                           {"03": ("03_1Q実績_2026.03", "col_prev"),
                            "08": (r"期間揃え比較\02_組替1Q_2025.05", "col_prev")})},
    },
    # 通常版の2025年3月期。当期列＝2026系列の前期列、前期列＝2024系列の当期列で
    # 組み直した 21_〜28_（build_std2025_series.py が生成）を読む。
    "fy2025": {
        "label": "2025年3月期",
        "curr_fy": {"03": "FY2025", "08": "FY2024"},
        "prev_fy": {"03": "FY2024", "08": "FY2023"},
        "pp_fy":   {"03": "FY2023", "08": "FY2022"},
        "periods": {
            "fy":    ("21_通期実績_2025.03", {"03": ("10_通期実績_2024.03", "col_prev"),
                                              "08": ("10_通期実績_2024.03", "col_prev")}),
            "q1":    ("22_1Q実績_2025.03",   {"03": ("11_1Q実績_2024.03", "col_prev"),
                                              "08": ("11_1Q実績_2024.03", "col_prev")}),
            "q2":    ("23_2Q単独_2025.03",   {"03": ("12_2Q単独_2024.03", "col_prev"),
                                              "08": ("12_2Q単独_2024.03", "col_prev")}),
            "h1":    ("24_上期実績_2025.03", {"03": ("13_上期実績_2024.03", "col_prev"),
                                              "08": ("13_上期実績_2024.03", "col_prev")}),
            "q3":    ("25_3Q単独_2025.03",   {"03": ("14_3Q単独_2024.03", "col_prev"),
                                              "08": ("14_3Q単独_2024.03", "col_prev")}),
            "q3cum": ("26_3Q累計_2025.03",   {"03": ("15_3Q累計_2024.03", "col_prev"),
                                              "08": ("15_3Q累計_2024.03", "col_prev")}),
            "q4":    ("27_4Q単独_2025.03",   {"03": ("16_4Q単独_2024.03", "col_prev"),
                                              "08": ("16_4Q単独_2024.03", "col_prev")}),
            "h2":    ("28_下期実績_2025.03", {"03": ("17_下期実績_2024.03", "col_prev"),
                                              "08": ("17_下期実績_2024.03", "col_prev")}),
        },
    },
    # 組替版の2025年3月期。期間揃え比較 09_〜16_ の当期列が
    # 「2024年3月〜2025年2月」＝3月決算社のFY2025にあたる。
    "fy2025_kumikae": {
        "label": "2025年3月期（組替）",
        "kumikae_base": 2025,   # 8月決算社の表示ラベルは実期間に置き換える
        "curr_fy": {"03": "FY2025", "08": "FY2025"},
        "prev_fy": {"03": "FY2024", "08": "FY2024"},
        "pp_fy":   {"03": "FY2023", "08": "FY2023"},
        "periods": {
            "fy":    (r"期間揃え比較\09_組替通期_2025.02", None),
            "q1":    (r"期間揃え比較\10_組替1Q_2024.05", None),
            "q2":    (r"期間揃え比較\11_組替2Q単独_2024.08", None),
            "h1":    (r"期間揃え比較\12_組替上期_2024.08", None),
            "q3":    (r"期間揃え比較\13_組替3Q単独_2024.11", None),
            "q3cum": (r"期間揃え比較\14_組替3Q累計_2024.11", None),
            "q4":    (r"期間揃え比較\15_組替4Q単独_2025.02", None),
            "h2":    (r"期間揃え比較\16_組替下期_2025.02", None),
        },
    },
    "fy2024": {
        "label": "2024年3月期",
        "curr_fy": {"03": "FY2024", "08": "FY2023"},
        "prev_fy": {"03": "FY2023", "08": "FY2022"},
        "pp_fy":   {"03": "FY2022", "08": "FY2021"},
        "periods": {
            "fy":    ("10_通期実績_2024.03", None),
            "q1":    ("11_1Q実績_2024.03", None),
            "q2":    ("12_2Q単独_2024.03", None),
            "h1":    ("13_上期実績_2024.03", None),
            "q3":    ("14_3Q単独_2024.03", None),
            "q3cum": ("15_3Q累計_2024.03", None),
            "q4":    ("16_4Q単独_2024.03", None),
            "h2":    ("17_下期実績_2024.03", None),
        },
    },
}
SUFFIX = {"fy": "", "q1": "-Q1", "q2": "-Q2", "h1": "-H1",
          "q3": "-Q3", "q3cum": "-Q3CUM", "q4": "-Q4", "h2": "-H2"}
# 8月決算社は暦月を合わせるため四半期番号がずれる年度がある。
# 組替版fy2027の1Q（3月決算社の2026年4-6月）に対し、ビック・コジマは2026年8月期3Q単独(3-5月)。
SUFFIX_OVERRIDE = {("fy2027_kumikae", "q1", "08"): "-Q3"}


def to_num(v):
    if v is None or (isinstance(v, str) and v.startswith("=")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_one(fy_key, conf, period, folders):
    curr_folder, pp_src = folders
    ws = data_sheet(load_workbook(find_latest_xlsx(PROJ / curr_folder), data_only=False))
    # 前々期は決算月ごとに取得元ブック・参照列が異なりうるのでここで解決する。
    pp_books = {}
    for fe, spec in (pp_src or {}).items():
        folder, col_key = spec
        pp_books[fe] = (data_sheet(load_workbook(find_latest_xlsx(PROJ / folder),
                                                 data_only=True)), col_key)

    companies_out = []
    for co in COMPANIES_MAIN:
        fe = co["fy_end"]
        ws_pp, pp_col = pp_books.get(fe, (None, "col_curr"))
        suf = SUFFIX_OVERRIDE.get((fy_key, period, fe), SUFFIX[period])
        # 表示用の実期間ラベル。組替版の8月決算社は自社の四半期番号がずれるため
        # kumikae_labels_08() で暦月に対応する実期間を割り当てる。
        base = conf.get("kumikae_base")
        if base and fe == "08":
            lc, lp, lpp = kumikae_labels_08(period, base)
        else:
            lc = period_label(conf["curr_fy"][fe] + suf, fe)
            lp = period_label(conf["prev_fy"][fe] + suf, fe)
            lpp = period_label(conf["pp_fy"][fe] + suf, fe)
        labels = {"current": lc, "previous": lp, "prev_previous": lpp}
        metrics = {}
        for row, key in ROW_TO_METRIC.items():
            metrics[key] = {
                "prev_previous": (to_num(ws_pp.cell(row=row, column=co[pp_col]).value)
                                  if ws_pp else None),
                "current": to_num(ws.cell(row=row, column=co["col_curr"]).value),
                "previous": to_num(ws.cell(row=row, column=co["col_prev"]).value),
                "forecast": None,
                "forecast_annual": None,
                "unit": "百万円",
            }

        yoy = {}
        for key in ("Revenue", "OperatingIncome", "OrdinaryIncome", "NetIncome"):
            m = metrics[key]
            cur, prev, pp = m["current"], m["previous"], m["prev_previous"]
            yoy[key] = {
                "current_yoy_pct": round((cur / prev - 1) * 100, 2) if cur and prev else None,
                "previous_yoy_pct": round((prev / pp - 1) * 100, 2) if prev and pp else None,
                "forecast_yoy_pct": None,
            }

        rev = metrics["Revenue"]["current"]
        rev_p = metrics["Revenue"]["previous"]
        rev_pp = metrics["Revenue"]["prev_previous"]
        ta, te = metrics["TotalAssets"]["current"], metrics["TotalEquity"]["current"]
        op, ord_c = metrics["OperatingIncome"]["current"], metrics["OrdinaryIncome"]["current"]
        ord_p, ord_pp = metrics["OrdinaryIncome"]["previous"], metrics["OrdinaryIncome"]["prev_previous"]
        gp_c, gp_p = metrics["GrossProfit"]["current"], metrics["GrossProfit"]["previous"]
        ratios = {
            "operating_margin_pct": margin(op, rev),
            "ordinary_margin_pct": margin(ord_c, rev),
            "ordinary_margin_pct_previous": margin(ord_p, rev_p),
            "ordinary_margin_pct_prev_previous": margin(ord_pp, rev_pp),
            "equity_ratio_pct": margin(te, ta),
            "gross_margin_pct": margin(gp_c, rev),
            "gross_margin_pct_previous": margin(gp_p, rev_p),
            "gross_margin_pct_prev_previous": margin(metrics["GrossProfit"]["prev_previous"], rev_pp),
            "net_margin_pct": margin(metrics["NetIncome"]["current"], rev),
            "net_margin_pct_previous": margin(metrics["NetIncome"]["previous"], rev_p),
            "net_margin_pct_prev_previous": margin(metrics["NetIncome"]["prev_previous"], rev_pp),
        }
        gm_c, gm_p = ratios["gross_margin_pct"], ratios["gross_margin_pct_previous"]
        ratios["gross_margin_pt_yoy"] = (round(gm_c - gm_p, 2)
                                         if gm_c is not None and gm_p is not None else None)
        om_c, om_p = ratios["ordinary_margin_pct"], ratios["ordinary_margin_pct_previous"]
        ratios["ordinary_margin_pt_yoy"] = (round(om_c - om_p, 2)
                                            if om_c is not None and om_p is not None else None)
        ratios["financial_leverage"] = round(ta / te, 3) if ta and te else None
        # 純利益率の前期差(pt)
        _nm_c, _nm_p = ratios["net_margin_pct"], ratios["net_margin_pct_previous"]
        ratios["net_margin_pt_yoy"] = (round(_nm_c - _nm_p, 2)
                                       if _nm_c is not None and _nm_p is not None else None)
        # 財務レバレッジの前期差(倍)。前期末BSは metrics の previous から算出する
        _ta_p, _te_p = metrics["TotalAssets"]["previous"], metrics["TotalEquity"]["previous"]
        _lev_p = round(_ta_p / _te_p, 3) if _ta_p and _te_p else None
        ratios["financial_leverage_previous"] = _lev_p
        ratios["financial_leverage_diff"] = (round(ratios["financial_leverage"] - _lev_p, 3)
                                             if ratios["financial_leverage"] is not None
                                             and _lev_p is not None else None)

        # 前々期ブックが無い年度は前々期のLTMを空にする（推移グラフの1点目が空欄になる）。
        ltm_data = (build_ltm(ws, ws_pp, co, pp_col) if ws_pp else
                    {"current": read_ltm(ws, co["col_curr"]),
                     "previous": read_ltm(ws, co["col_prev"]),
                     "prev_previous": read_ltm(None, None)})
        ratios["roe_pct"] = ltm_data["current"]["roe_pct"]
        ratios["roic_pct"] = ltm_data["current"]["roic_pct"]
        # 直近12ヶ月ベース指標の前期差
        def _diff(cur, prev, nd):
            return round(cur - prev, nd) if cur is not None and prev is not None else None
        _lp, _lc = ltm_data["previous"], ltm_data["current"]
        ratios["roe_pt_yoy"] = _diff(_lc["roe_pct"], _lp["roe_pct"], 2)
        ratios["roic_pt_yoy"] = _diff(_lc["roic_pct"], _lp["roic_pct"], 2)
        ratios["asset_turnover_diff"] = _diff(_lc["asset_turnover"], _lp["asset_turnover"], 3)
        ratios["inventory_turnover_ltm"] = _lc["inventory_turnover"]
        ratios["inventory_turnover_diff"] = _diff(_lc["inventory_turnover"], _lp["inventory_turnover"], 3)

        trend = {}
        for k in ("roe", "roic", "asset_turnover", "inventory_turnover", "gross_margin", "net_margin"):
            trend[k] = {"prev_previous": None, "previous": None, "current": None, "forecast": None}
        for p in ("prev_previous", "previous", "current"):
            trend["gross_margin"][p] = margin(metrics["GrossProfit"][p], metrics["Revenue"][p])
            trend["net_margin"][p] = margin(metrics["NetIncome"][p], metrics["Revenue"][p])
        trend["ltm_asset_turnover"] = ltm_trend(ltm_data, "asset_turnover")
        trend["ltm_inventory_turnover"] = ltm_trend(ltm_data, "inventory_turnover")
        trend["roe"] = ltm_trend(ltm_data, "roe_pct")
        trend["roic"] = ltm_trend(ltm_data, "roic_pct")

        if co.get("is_segment"):
            ratios["equity_ratio_pct"] = None

        companies_out.append({
            "key": co["key"], "label": co["label"], "ticker": co["ticker"],
            "fiscal_year_end": fe, "consolidation": co["consol"],
            "current_period": conf["curr_fy"][fe] + suf,
            "previous_period": conf["prev_fy"][fe] + suf,
            "prev_previous_period": conf["pp_fy"][fe] + suf,
            "forecast_period": None,
            "period_labels": labels,
            "tanshin_url": co["url"], "metrics": metrics, "yoy": yoy, "ratios": ratios,
            "trend_ratios": trend, "ltm": ltm_data,
            "is_segment": bool(co.get("is_segment")),
        })

    out = {
        "schema_version": "1.0",
        "generated_at": date.today().isoformat(),
        "source": f"各社決算短信 ({conf['label']} {period})",
        "period_type": f"{fy_key}_{period}",
        "companies": companies_out,
    }
    path = OUT_DIR / f"companies_{fy_key}_{period}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    filled = sum(1 for c in companies_out if c["metrics"]["Revenue"]["current"] is not None)
    print(f"  → {path.name}  （売上高が入っている社: {filled}/8）")


def main():
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    for fy_key, conf in FISCAL_YEARS.items():
        print(f"\n========== {fy_key} ({conf['label']}) ==========")
        for period, folders in conf["periods"].items():
            build_one(fy_key, conf, period, folders)


if __name__ == "__main__":
    main()
