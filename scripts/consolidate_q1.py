"""1Q実績版ダッシュボード用 companies_q1.json を生成。

データソース:
  当期(2026/3 1Q or 2025/8 1Q):  03_1Q実績_2026.03/_ver.7.xlsx
  前期(2025/3 1Q or 2024/8 1Q):  同上 (E列)
  前々期(2024/3 1Q or 2023/8 1Q): 11_1Q実績_2024.03/_ver.5.xlsx (D列)
  次期計画(通期):  consolidate.py と同じ FORECAST 固定値

BS時点:
  1Q短信のBSは「当1Q末 vs 前期末」なので、ratios は四半期PL/四半期末BS の組み合わせ。
  ROE/ROIC/総資産回転率は決算短信上の単純比率 (年率換算しない)。
"""
import json
from datetime import date
from pathlib import Path
from openpyxl import load_workbook

from xlsx_utils import find_latest_xlsx, data_sheet, build_ltm, ltm_trend, margin

ROOT = Path(__file__).parent.parent.parent

# 1Q Excelファイル: フォルダ内の最新ver（_ver.N 最大）を自動採用
XLSX = find_latest_xlsx(ROOT.parent / "03_1Q実績_2026.03")
PREV_PREV_XLSX = find_latest_xlsx(ROOT.parent / "11_1Q実績_2024.03")

OUT_DIR = Path(__file__).parent.parent / "data"
OUT_DIR.mkdir(exist_ok=True)

# 各社の Excel 列レイアウト
COMPANIES_MAIN = [
    {
        "key": "yamada", "label": "ヤマダHD", "ticker": "9831",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026-Q1", "previous_period": "FY2025-Q1",
        "prev_previous_period": "FY2024-Q1", "forecast_period": "FY2027",
        "col_curr": 4, "col_prev": 5,
        "tanshin_url": "https://www.yamada-holdings.jp/ir/",
    },
    {
        "key": "yamada_denki", "label": "ヤマダ（デンキセグメント）", "ticker": "9831-DK",
        "fiscal_year_end": "03", "consolidation": "segment",
        "current_period": "FY2026-Q1", "previous_period": "FY2025-Q1",
        "prev_previous_period": "FY2024-Q1", "forecast_period": "FY2027",
        "col_curr": 6, "col_prev": 7,
        "tanshin_url": "https://www.yamada-holdings.jp/ir/",
        "is_segment": True,
    },
    {
        "key": "ks", "label": "ケーズHD", "ticker": "8282",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026-Q1", "previous_period": "FY2025-Q1",
        "prev_previous_period": "FY2024-Q1", "forecast_period": "FY2027",
        "col_curr": 8, "col_prev": 9,
        "tanshin_url": "https://www.ksdenki.co.jp/ir/",
    },
    {
        "key": "edion", "label": "エディオン", "ticker": "2730",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026-Q1", "previous_period": "FY2025-Q1",
        "prev_previous_period": "FY2024-Q1", "forecast_period": "FY2027",
        "col_curr": 10, "col_prev": 11,
        "tanshin_url": "https://www.edion.co.jp/ir/",
    },
    {
        "key": "joshin", "label": "上新電機", "ticker": "8173",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026-Q1", "previous_period": "FY2025-Q1",
        "prev_previous_period": "FY2024-Q1", "forecast_period": "FY2027",
        "col_curr": 12, "col_prev": 13,
        "tanshin_url": "https://www.joshin.co.jp/ir/",
    },
    {
        "key": "nojima", "label": "ノジマ", "ticker": "7419",
        "fiscal_year_end": "03", "consolidation": "consolidated",
        "current_period": "FY2026-Q1", "previous_period": "FY2025-Q1",
        "prev_previous_period": "FY2024-Q1", "forecast_period": "FY2027",
        "col_curr": 14, "col_prev": 15,
        "tanshin_url": "https://www.nojima.co.jp/ir/",
    },
    {
        "key": "bic", "label": "ビックカメラ（連結）", "ticker": "3048",
        "fiscal_year_end": "08", "consolidation": "consolidated",
        "current_period": "FY2025-Q1", "previous_period": "FY2024-Q1",
        "prev_previous_period": "FY2023-Q1", "forecast_period": "FY2026",
        "col_curr": 22, "col_prev": 23,
        "tanshin_url": "https://www.biccamera.co.jp/ir/",
    },
    {
        "key": "kojima", "label": "コジマ", "ticker": "7513",
        "fiscal_year_end": "08", "consolidation": "non_consolidated",
        "current_period": "FY2025-Q1", "previous_period": "FY2024-Q1",
        "prev_previous_period": "FY2023-Q1", "forecast_period": "FY2026",
        "col_curr": 18, "col_prev": 19,
        "tanshin_url": "https://www.kojima.net/corporation/ir/",
    },
]

ROW_TO_METRIC = {
    7:  "Revenue",
    8:  "GrossProfit",
    9:  "SGA",
    82: "OperatingIncome",
    83: "OrdinaryIncome",
    84: "NetIncome",
    88: "InterestBearingDebt",
    89: "TotalEquity",
    100: "TotalAssets",
    101: "Inventory",
    86: "Tax",
}

# 通期計画予想値 (consolidate.pyと同じ。1Qダッシュボード末尾の「次期通期業績予想」テーブル用)
FORECAST = {
    "yamada":       {"Revenue": 1780000, "OperatingIncome": 51500, "OrdinaryIncome": 52600, "NetIncome": 27800, "GrossProfit": 503100},
    "ks":           {"Revenue": 785000,  "OperatingIncome": 30500, "OrdinaryIncome": 33500, "NetIncome": 20000, "GrossProfit": 219500},
    "edion":        {"Revenue": 816000,  "OperatingIncome": 27000, "OrdinaryIncome": 27000, "NetIncome": 15700, "GrossProfit": 235800},
    "joshin":       {"Revenue": 438000,  "OperatingIncome": 6000,  "OrdinaryIncome": 5500,  "NetIncome": 3500,  "GrossProfit": 114500},
    "nojima":       {"Revenue": 1000000, "OperatingIncome": 59000, "OrdinaryIncome": 76000, "NetIncome": 48000, "GrossProfit": None},
    "bic":          {"Revenue": 1013000, "OperatingIncome": 30500, "OrdinaryIncome": 31500, "NetIncome": 17500, "GrossProfit": 271484},
    "kojima":       {"Revenue": 294000,  "OperatingIncome": 7600,  "OrdinaryIncome": 7900,  "NetIncome": 4900,  "GrossProfit": 79968},
    "yamada_denki": {"Revenue": 1407400, "OperatingIncome": 34500, "OrdinaryIncome": 37100, "NetIncome": None,  "GrossProfit": 411500},
}


def to_num(v):
    if v is None or (isinstance(v, str) and v.startswith("=")):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    wb = load_workbook(XLSX, data_only=False)
    ws = data_sheet(wb)
    print(f"sheet: {ws.title}")
    wb_pp = load_workbook(PREV_PREV_XLSX, data_only=True)
    ws_pp = data_sheet(wb_pp)

    companies_out = []
    for co in COMPANIES_MAIN:
        metrics = {}
        for row, key in ROW_TO_METRIC.items():
            cur = to_num(ws.cell(row=row, column=co["col_curr"]).value)
            prev = to_num(ws.cell(row=row, column=co["col_prev"]).value)
            prev_prev = to_num(ws_pp.cell(row=row, column=co["col_curr"]).value)
            fc_annual = FORECAST.get(co["key"], {}).get(key)  # 通期計画値（1Qの4倍ではなく通期）
            metrics[key] = {
                "prev_previous": prev_prev,
                "current": cur,
                "previous": prev,
                # forecast: 1Q推移グラフ用 → null (4点目を描画しない=スケール歪み回避)
                "forecast": None,
                # forecast_annual: 次期通期業績予想テーブル用 (1Qでも通期計画は表示)
                "forecast_annual": fc_annual,
                "unit": "百万円",
            }

        # YoY (current vs previous, previous vs prev_previous), forecast_yoy_pct は1Qでは null
        yoy = {}
        for key in ("Revenue", "OperatingIncome", "OrdinaryIncome", "NetIncome"):
            m = metrics.get(key, {})
            cur = m.get("current")
            prev = m.get("previous")
            pp = m.get("prev_previous")
            yoy[key] = {
                "current_yoy_pct": round((cur / prev - 1) * 100, 2) if cur and prev else None,
                "previous_yoy_pct": round((prev / pp - 1) * 100, 2) if prev and pp else None,
                "forecast_yoy_pct": None,
            }

        # 当期スポット比率
        rev = metrics["Revenue"]["current"]
        op = metrics["OperatingIncome"]["current"]
        ta = metrics["TotalAssets"]["current"]
        te = metrics["TotalEquity"]["current"]
        ord_curr = metrics["OrdinaryIncome"]["current"]
        rev_prev = metrics["Revenue"]["previous"]
        ord_prev = metrics["OrdinaryIncome"]["previous"]
        rev_pp = metrics["Revenue"]["prev_previous"]
        ord_pp = metrics["OrdinaryIncome"]["prev_previous"]
        gp_curr = metrics["GrossProfit"]["current"]
        gp_prev = metrics["GrossProfit"]["previous"]
        gp_pp = metrics["GrossProfit"]["prev_previous"]
        ratios = {
            "operating_margin_pct": round(op / rev * 100, 2) if op and rev else None,
            "ordinary_margin_pct": round(ord_curr / rev * 100, 2) if ord_curr and rev else None,
            "ordinary_margin_pct_previous": round(ord_prev / rev_prev * 100, 2) if ord_prev and rev_prev else None,
            "ordinary_margin_pct_prev_previous": round(ord_pp / rev_pp * 100, 2) if ord_pp and rev_pp else None,
            "equity_ratio_pct": round(te / ta * 100, 2) if te and ta else None,
            "gross_margin_pct": margin(gp_curr, rev),
            "gross_margin_pct_previous": margin(gp_prev, rev_prev),
            "gross_margin_pct_prev_previous": margin(gp_pp, rev_pp),
            "gross_margin_pt_yoy": (round(margin(gp_curr, rev) - margin(gp_prev, rev_prev), 2)
                                    if margin(gp_curr, rev) is not None
                                    and margin(gp_prev, rev_prev) is not None else None),
            "net_margin_pct": margin(metrics["NetIncome"]["current"], rev),
            "net_margin_pct_previous": margin(metrics["NetIncome"]["previous"], rev_prev),
            "net_margin_pct_prev_previous": margin(metrics["NetIncome"]["prev_previous"], rev_pp),
        }

        # 経常利益率の前期差(pt)、財務レバレッジ、ROE
        # 財務レバレッジ・ROEはExcelのR104(=総資産÷自己資本)・R98(=純利益÷自己資本×100)と同じ定義。
        # デンキセグメント等はBS非開示のため自己資本がなく、いずれもNoneになる。
        _om_c, _om_p = ratios["ordinary_margin_pct"], ratios["ordinary_margin_pct_previous"]
        ratios["ordinary_margin_pt_yoy"] = (round(_om_c - _om_p, 2)
                                            if _om_c is not None and _om_p is not None else None)
        ratios["financial_leverage"] = round(ta / te, 3) if ta and te else None
        # 純利益率の前期差(pt)
        _nm_c, _nm_p = ratios["net_margin_pct"], ratios["net_margin_pct_previous"]
        ratios["net_margin_pt_yoy"] = (round(_nm_c - _nm_p, 2)
                                       if _nm_c is not None and _nm_p is not None else None)
        # 財務レバレッジの前期差(倍)。前期末BSは metrics の previous から算出する
        _ta_p, _te_p = metrics["TotalAssets"]["previous"], metrics["TotalEquity"]["previous"]
        _lev_p = round(_ta_p / _te_p, 3) if _ta_p and _te_p else None
        ratios["financial_leverage_previous"] = _lev_p
        ratios["financial_leverage_diff"] = (round(ratios["financial_leverage"] - _lev_p, 3)
                                             if ratios["financial_leverage"] is not None
                                             and _lev_p is not None else None)
        ratios["roe_pct"] = margin(metrics["NetIncome"]["current"], te)

        # 推移用比率 (前々期/前期/当期、1Qのforecastは省略=null表示)
        def safe_div(a, b):
            try:
                return a / b if a is not None and b not in (None, 0) else None
            except (TypeError, ZeroDivisionError):
                return None

        def roe(ni, eq):
            v = safe_div(ni, eq)
            return round(v * 100, 2) if v is not None else None

        EFFECTIVE_TAX_RATE = 0.35

        def roic(op_v, debt_v, eq_v):
            ic = (debt_v or 0) + (eq_v or 0)
            if not op_v or ic == 0:
                return None
            nopat = op_v * (1 - EFFECTIVE_TAX_RATE)
            return round(nopat / ic * 100, 2)

        def turnover(num, den):
            v = safe_div(num, den)
            return round(v, 3) if v is not None else None

        bs_periods = {
            "prev_previous": {
                "ta": metrics["TotalAssets"]["prev_previous"],
                "te": metrics["TotalEquity"]["prev_previous"],
                "inv": metrics["Inventory"]["prev_previous"],
                "debt": metrics["InterestBearingDebt"]["prev_previous"],
            },
            "previous": {
                "ta": metrics["TotalAssets"]["previous"],
                "te": metrics["TotalEquity"]["previous"],
                "inv": metrics["Inventory"]["previous"],
                "debt": metrics["InterestBearingDebt"]["previous"],
            },
            "current": {
                "ta": metrics["TotalAssets"]["current"],
                "te": metrics["TotalEquity"]["current"],
                "inv": metrics["Inventory"]["current"],
                "debt": metrics["InterestBearingDebt"]["current"],
            },
        }

        trend_ratios = {"roe": {}, "roic": {}, "asset_turnover": {},
                        "inventory_turnover": {}, "gross_margin": {}, "net_margin": {}}
        for period in ("prev_previous", "previous", "current"):
            rev_p = metrics["Revenue"][period]
            op_p = metrics["OperatingIncome"][period]
            ni_p = metrics["NetIncome"][period]
            bs = bs_periods[period]
            trend_ratios["roe"][period] = roe(ni_p, bs["te"])
            trend_ratios["roic"][period] = roic(op_p, bs["debt"], bs["te"])
            trend_ratios["asset_turnover"][period] = turnover(rev_p, bs["ta"])
            trend_ratios["inventory_turnover"][period] = turnover(rev_p, bs["inv"])
            trend_ratios["gross_margin"][period] = margin(metrics["GrossProfit"][period], rev_p)
            trend_ratios["net_margin"][period] = margin(ni_p, rev_p)
        # forecast は1Qでは null (年間ベースの比率と1Qベースの比率は単純比較不能)
        for k in ("roe", "roic", "asset_turnover", "inventory_turnover",
                  "gross_margin", "net_margin"):
            trend_ratios[k]["forecast"] = None

        # ヤマダデンキセグ: BS非開示 → ROE/ROIC/総資産回転率 null
        if co.get("is_segment"):
            for k in ("roe", "roic", "asset_turnover"):
                trend_ratios[k] = {"prev_previous": None, "previous": None, "current": None, "forecast": None}
            # 在庫回転率は POSベース固定値 (1Qダッシュボードでは年率換算しない=年率値そのまま表示)
            trend_ratios["inventory_turnover"] = {
                "prev_previous": 3.6, "previous": 4.0, "current": 4.5, "forecast": 5.0,
            }
            trend_ratios["inventory_turnover_override"] = {
                "basis": "デンキセグメント（ヤマダデンキPOSベース）",
                "source": "ヤマダ_決算説明会資料 (年率換算前POSベース)",
                "note": "1Qダッシュボードでも年率POSベース値を流用表示",
            }
            ratios["equity_ratio_pct"] = None

        # 直近四半期(LTM)回転率も推移グラフ用の形に持たせる
        ltm_data = build_ltm(ws, ws_pp, co)
        trend_ratios["ltm_asset_turnover"] = ltm_trend(ltm_data, "asset_turnover")
        trend_ratios["ltm_inventory_turnover"] = ltm_trend(ltm_data, "inventory_turnover")
        # ROE / ROIC は直近12ヶ月ベースに統一する。
        # 四半期の利益をそのまま自己資本で割ると期間の短さだけで小さく出るため。
        # 次期予想の点は従来どおり計画値ベースで残す（LTMは実績のみ算定できるため）
        _fc_roe = trend_ratios.get("roe", {}).get("forecast")
        _fc_roic = trend_ratios.get("roic", {}).get("forecast")
        trend_ratios["roe"] = ltm_trend(ltm_data, "roe_pct")
        trend_ratios["roic"] = ltm_trend(ltm_data, "roic_pct")
        trend_ratios["roe"]["forecast"] = _fc_roe
        trend_ratios["roic"]["forecast"] = _fc_roic
        ratios["roe_pct"] = ltm_data["current"]["roe_pct"]
        ratios["roic_pct"] = ltm_data["current"]["roic_pct"]
        # 直近12ヶ月ベース指標の前期差
        def _diff(cur, prev, nd):
            return round(cur - prev, nd) if cur is not None and prev is not None else None
        _lp, _lc = ltm_data["previous"], ltm_data["current"]
        ratios["roe_pt_yoy"] = _diff(_lc["roe_pct"], _lp["roe_pct"], 2)
        ratios["roic_pt_yoy"] = _diff(_lc["roic_pct"], _lp["roic_pct"], 2)
        ratios["asset_turnover_diff"] = _diff(_lc["asset_turnover"], _lp["asset_turnover"], 3)

        co_out = {
            "key": co["key"],
            "label": co["label"],
            "ticker": co["ticker"],
            "fiscal_year_end": co["fiscal_year_end"],
            "consolidation": co["consolidation"],
            "current_period": co["current_period"],
            "previous_period": co["previous_period"],
            "prev_previous_period": co["prev_previous_period"],
            "forecast_period": co["forecast_period"],
            "tanshin_url": co["tanshin_url"],
            "metrics": metrics,
            "yoy": yoy,
            "ratios": ratios,
            "trend_ratios": trend_ratios,
            "ltm": ltm_data,
            "is_segment": bool(co.get("is_segment")),
        }
        companies_out.append(co_out)

    output = {
        "schema_version": "1.0",
        "generated_at": date.today().isoformat(),
        "source": "各社決算短信・決算説明会資料 (TDnet公開資料、2026/8月発表分の1Q決算)",
        "note": "ヤマダ・ケーズ・エディオン・上新・ノジマ=2026年3月期第1四半期。ビックカメラ・コジマ=2025年8月期第1四半期。",
        "period_type": "quarterly_1Q",
        "companies": companies_out,
    }

    out_path = OUT_DIR / "companies_q1.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path}")
    print(f"Companies: {len(companies_out)}")
    for c in companies_out:
        rev = c["metrics"]["Revenue"]["current"]
        op = c["metrics"]["OperatingIncome"]["current"]
        om = c["ratios"]["operating_margin_pct"]
        print(f"  {c['label']:20s} ({c['current_period']}): 売上 {rev:>10,.0f} / 営利 {op:>7,.0f} / 営利率 {om}%")


if __name__ == "__main__":
    main()
